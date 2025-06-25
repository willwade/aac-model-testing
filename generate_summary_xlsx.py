#!/usr/bin/env python3
"""
Generate Excel Summary from AAC Model Testing Results

This script creates a comprehensive Excel spreadsheet from all AAC model test results,
including detailed test responses, metrics, and device information.

Usage:
    python generate_summary_xlsx.py                    # Generate summary.xlsx
    python generate_summary_xlsx.py --output my_report.xlsx  # Custom filename
    python generate_summary_xlsx.py --device work-laptop     # Filter by device
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def load_results_history(results_dir: Path) -> List[Dict[str, Any]]:
    """Load all results from the master results file."""
    master_file = results_dir / "all_results.jsonl"
    
    if not master_file.exists():
        print(f"No results file found at {master_file}")
        return []
    
    results = []
    try:
        with open(master_file, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    results.append(json.loads(line))
    except Exception as e:
        print(f"Error loading results history: {e}")
        return []
    
    return results


def load_raw_results(results_dir: Path, results_file: str) -> Optional[Dict[str, Any]]:
    """Load detailed raw results from a specific results file."""
    raw_file = results_dir / results_file
    
    if not raw_file.exists():
        print(f"Warning: Raw results file not found: {raw_file}")
        return None
    
    try:
        with open(raw_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading raw results from {raw_file}: {e}")
        return None


def format_timestamp(timestamp_str: str) -> str:
    """Format timestamp for display."""
    try:
        dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except:
        return timestamp_str


def create_summary_sheet(workbook: Workbook, history: List[Dict[str, Any]]) -> None:
    """Create the main summary sheet with test run overview."""
    ws = workbook.active
    ws.title = "Test Run Summary"
    
    # Headers
    headers = [
        'Timestamp', 'Device', 'Models Tested', 'Test Cases', 
        'Best Model', 'Best Score', 'Avg Response Time (s)', 'Avg Memory (MB)'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, entry in enumerate(history, 2):
        ws.cell(row=row, column=1, value=format_timestamp(entry['timestamp']))
        ws.cell(row=row, column=2, value=entry.get('device_name', 'unknown'))
        ws.cell(row=row, column=3, value=', '.join(entry['models']))
        ws.cell(row=row, column=4, value=', '.join(entry['test_cases']))
        
        # Find best model for this run
        best_model = None
        best_score = 0
        avg_time = 0
        avg_memory = 0
        model_count = 0
        
        for model, summary in entry.get('summary', {}).items():
            if summary and 'overall_score' in summary:
                score = summary['overall_score']
                if score > best_score:
                    best_score = score
                    best_model = model
                avg_time += summary.get('average_response_time', 0)
                avg_memory += summary.get('average_memory_usage', 0)
                model_count += 1
        
        if model_count > 0:
            avg_time /= model_count
            avg_memory /= model_count
        
        ws.cell(row=row, column=5, value=best_model or 'N/A')
        ws.cell(row=row, column=6, value=round(best_score, 4) if best_score > 0 else 'N/A')
        ws.cell(row=row, column=7, value=round(avg_time, 2) if avg_time > 0 else 'N/A')
        ws.cell(row=row, column=8, value=round(avg_memory, 1) if avg_memory > 0 else 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_model_performance_sheet(workbook: Workbook, history: List[Dict[str, Any]]) -> None:
    """Create a detailed model performance comparison sheet."""
    ws = workbook.create_sheet("Model Performance")
    
    # Collect all model performance data
    performance_data = []
    
    for entry in history:
        timestamp = format_timestamp(entry['timestamp'])
        device = entry.get('device_name', 'unknown')
        
        for model, summary in entry.get('summary', {}).items():
            if summary:
                performance_data.append({
                    'Timestamp': timestamp,
                    'Device': device,
                    'Model': model,
                    'Total Tests': summary.get('total_tests', 0),
                    'Successful Tests': summary.get('successful_tests', 0),
                    'Success Rate (%)': round((summary.get('successful_tests', 0) / max(summary.get('total_tests', 1), 1)) * 100, 1),
                    'Overall Score': round(summary.get('overall_score', 0), 4),
                    'Avg Response Time (s)': round(summary.get('average_response_time', 0), 2),
                    'Avg Memory Usage (MB)': round(summary.get('average_memory_usage', 0), 1)
                })
    
    # Convert to DataFrame and write to sheet
    if performance_data:
        df = pd.DataFrame(performance_data)
        
        # Add headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


def create_detailed_responses_sheet(workbook: Workbook, history: List[Dict[str, Any]], results_dir: Path) -> None:
    """Create a sheet with all detailed test responses."""
    ws = workbook.create_sheet("Detailed Responses")
    
    # Headers for detailed responses
    headers = [
        'Timestamp', 'Device', 'Model', 'Test Case', 'Test ID', 'Input Text', 
        'Model Response', 'Response Time (s)', 'Score', 'Passed', 'Feedback',
        'Grammar Score', 'Completeness Score', 'Naturalness Score', 'Semantic Preservation'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    current_row = 2
    
    # Process each test run
    for entry in history:
        timestamp = format_timestamp(entry['timestamp'])
        device = entry.get('device_name', 'unknown')
        results_file = entry.get('results_file')
        
        if not results_file:
            continue
            
        # Load raw results
        raw_data = load_raw_results(results_dir, results_file)
        if not raw_data:
            continue
        
        # Extract detailed responses
        for model_name, model_data in raw_data.get('results', {}).items():
            test_results = model_data.get('test_results', {})
            
            for test_case, test_data in test_results.items():
                individual_results = test_data.get('individual_results', [])
                
                for result in individual_results:
                    # Add row data
                    row_data = [
                        timestamp,
                        device,
                        model_name,
                        test_case,
                        result.get('test_item_id', ''),
                        result.get('input', ''),
                        result.get('response', ''),
                        round(result.get('response_time', 0), 3),
                        round(result.get('score', 0), 4),
                        'Yes' if result.get('passed', False) else 'No',
                        result.get('feedback', ''),
                        round(result.get('metrics', {}).get('grammar_score', 0), 3),
                        round(result.get('metrics', {}).get('completeness_score', 0), 3),
                        round(result.get('metrics', {}).get('naturalness_score', 0), 3),
                        round(result.get('metrics', {}).get('semantic_preservation', 0), 3)
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        # Wrap text for response column
                        if col == 7:  # Response column
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    current_row += 1
    
    # Auto-adjust column widths (with special handling for response column)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Special handling for response column (column G)
        if column_letter == 'G':
            ws.column_dimensions[column_letter].width = 50
            continue
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_device_info_sheet(workbook: Workbook, history: List[Dict[str, Any]], results_dir: Path) -> None:
    """Create a sheet with device information from test runs."""
    ws = workbook.create_sheet("Device Information")
    
    # Collect unique device info
    device_info_data = {}
    
    for entry in history:
        results_file = entry.get('results_file')
        if not results_file:
            continue
            
        raw_data = load_raw_results(results_dir, results_file)
        if not raw_data:
            continue
        
        device_info = raw_data.get('metadata', {}).get('device_info', {})
        device_name = device_info.get('device_name', 'unknown')
        
        if device_name not in device_info_data:
            device_info_data[device_name] = device_info
    
    # Headers
    headers = [
        'Device Name', 'Hostname', 'Platform', 'System', 'Processor', 
        'Architecture', 'Python Version', 'CPU Count', 'Logical CPUs',
        'Total Memory (GB)', 'Available Memory (GB)'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Add device data
    for row, (device_name, info) in enumerate(device_info_data.items(), 2):
        row_data = [
            info.get('device_name', ''),
            info.get('hostname', ''),
            info.get('platform', ''),
            info.get('system', ''),
            info.get('processor', ''),
            info.get('architecture', ''),
            info.get('python_version', ''),
            info.get('cpu_count', ''),
            info.get('cpu_count_logical', ''),
            info.get('total_memory_gb', ''),
            info.get('available_memory_gb', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Generate comprehensive Excel summary from AAC model testing results",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_summary_xlsx.py                    # Generate summary.xlsx
  python generate_summary_xlsx.py --output report.xlsx  # Custom filename
  python generate_summary_xlsx.py --device work-laptop  # Filter by device
        """
    )
    
    parser.add_argument(
        "--results-dir",
        type=str,
        default="results",
        help="Results directory (default: results)"
    )
    
    parser.add_argument(
        "--output",
        type=str,
        default="aac_model_testing_summary.xlsx",
        help="Output Excel filename (default: aac_model_testing_summary.xlsx)"
    )
    
    parser.add_argument(
        "--device",
        type=str,
        help="Filter results by device name"
    )
    
    args = parser.parse_args()
    
    results_dir = Path(args.results_dir)
    
    # Load results history
    print("Loading results history...")
    history = load_results_history(results_dir)
    
    if not history:
        print("No test results found. Run some tests first!")
        return
    
    # Filter by device if specified
    if args.device:
        history = [h for h in history if h.get('device_name', '').lower() == args.device.lower()]
        if not history:
            print(f"No results found for device: {args.device}")
            return
        print(f"Filtered to {len(history)} results for device: {args.device}")
    
    print(f"Processing {len(history)} test runs...")
    
    # Create Excel workbook
    workbook = Workbook()
    
    try:
        # Create all sheets
        print("Creating summary sheet...")
        create_summary_sheet(workbook, history)
        
        print("Creating model performance sheet...")
        create_model_performance_sheet(workbook, history)
        
        print("Creating detailed responses sheet...")
        create_detailed_responses_sheet(workbook, history, results_dir)
        
        print("Creating device information sheet...")
        create_device_info_sheet(workbook, history, results_dir)
        
        # Save the workbook
        output_path = Path(args.output)
        workbook.save(output_path)
        
        print(f"\n✅ Excel summary generated successfully!")
        print(f"📁 File saved as: {output_path.absolute()}")
        print(f"📊 Contains {len(history)} test runs with detailed analysis")
        
        # Print summary stats
        total_models = set()
        total_tests = 0
        for entry in history:
            total_models.update(entry['models'])
            for model, summary in entry.get('summary', {}).items():
                if summary:
                    total_tests += summary.get('total_tests', 0)
        
        print(f"🤖 Models tested: {len(total_models)} ({', '.join(sorted(total_models))})")
        print(f"🧪 Total individual tests: {total_tests}")
        
    except Exception as e:
        print(f"❌ Error generating Excel file: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
